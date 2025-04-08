import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

print("Script started")

# Directory containing the log files
log_dir = "/projects/digimind/orig/Behavioral/presentation logs/video_task"

# Template Excel file
template_file = "/projects/digimind/arna/digimind_video_task_experience_template.xlsx"

# Output directory for updated Excel files
output_dir = "/projects/digimind/arna/beh"

import re

def collect_response_codes_from_file(log_file):
    # List to store the response codes
    response_numbers = []
    subject_number = None  # To store the subject number

    with open(log_file, 'r') as file:
        lines = file.readlines()
        for line in lines:
            # Split the line into words
            words = line.split()
            
            # Check if 'Picture' is in the line and if it's a valid picture (not 'break')
            if 'Picture' in words and 'break' not in words:
                picture = words[2]  # e.g., 'q1', 'q2', etc.
                
                # Store the subject number (assumed to be in the first column of the first row)
                if subject_number is None:
                    subject_number = words[0]
                
                # Find the next picture after the current one
                next_picture = None
                for next_line in lines[lines.index(line) + 1:]:
                    next_words = next_line.split()
                    if 'Picture' in next_words and 'break' not in next_words:
                        next_picture = next_words[2]  # e.g., 'q2', 'q3', etc.
                        break
                
                # If the next picture exists, check the responses between them
                if next_picture:
                    responses = []
                    for check_line in lines[lines.index(line) + 1:lines.index(next_line)]:
                        check_words = check_line.split()
                        if 'Response' in check_words:
                            response_code = check_words[3]  # Assume the code is in the 4th column
                            if response_code.isdigit():
                                responses.append(response_code)
                    # If there are multiple responses, take the latest one
                    if len(responses) > 1:
                        if responses[-1] == '9':
                            response_numbers.append(responses[-2])
                        else:
                            response_numbers.append(responses[-1])  # Take the last response

                    # If there are two responses, check if the second one is 9
                    #if len(responses) > 1:
                     #   if responses[1] == '9':
                      #      response_numbers.append(responses[0])  # Take the first response
                       # else:
                        #    response_numbers.append(responses[1])  # Take the second response
                    elif responses:
                        response_numbers.append(responses[0])  # Take the first response if only one exists
                        
    # Check if the total number of responses is higher than 225
    if len(response_numbers) != 225:
        print(f"Error: Total number of responses is {len(response_numbers)}, which differs from 225. Subject: {subject_number}")

    return response_numbers


def update_excel(template_path, output_path, subject_id, response_data):
    """Update the Excel template with the subject ID and response data."""
    # Load the Excel template
    workbook = load_workbook(template_path)
    sheet = workbook["time"]

    # Update B1 with subject ID
    sheet["B1"] = subject_id[1:]


    # Map responses to the appropriate cells (C18 to AA26)
# Generate cell references filling column-wise from C18 to AA26
    response_cells = [
        f"{get_column_letter(col)}{row}" for col in range(3, 29) for row in range(18, 27)
    ]

    # Ensure all responses are numbers (convert to int or float)
    for i, response in enumerate(response_data):
        if i < len(response_cells):
            try:
                # Try converting the response to a number
                sheet[response_cells[i]] = float(response) if '.' in response else int(response)
            except ValueError:
                # If conversion fails, you can either set it to None or keep it as string
                sheet[response_cells[i]] = response

    # Save the updated file
    workbook.save(output_path)

def process_log_files(log_dir, template_file, output_dir):
    """Process all log files and update Excel templates."""
    print('process')
    total_files = 0
    files_with_error = 0
    for log_file in os.listdir(log_dir):
        if log_file.endswith("-video_task.log"):
            subject_id = log_file.split("-")[0]
            log_path = os.path.join(log_dir, log_file)
            response_data = collect_response_codes_from_file(log_path)

            #print(response_data)
            if response_data:
                output_file = os.path.join(output_dir, f"{subject_id}_video_task_experience.xlsx")
                update_excel(template_file, output_file, subject_id, response_data)
                print(f"Processed: {log_file} -> {output_file}")

# Execute the script
print('function')
process_log_files(log_dir, template_file, output_dir)
