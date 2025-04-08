import openpyxl
import os

def copy_categorized_values(source_file, target_file):
    # Load the source workbook and 'categorized' sheet
    source_wb = openpyxl.load_workbook(source_file, data_only=True)  # data_only ensures formulas are evaluated
    categorized_sheet = source_wb['categorized']

    # Read the identifier from B1 of the source file (subject ID)
    subject_id = categorized_sheet['B1'].value
    print(f"Processing Subject ID: {subject_id}")

    # Load the target workbook
    target_wb = openpyxl.load_workbook(target_file)

    # Loop through q1_cat to q9_cat sheets in the target file
    for sheet_index in range(1, 10):
        sheet_name = f"q{sheet_index}_cat"
        target_sheet = target_wb[sheet_name]
        print(f"Processing sheet: {sheet_name}")

        # Search for the matching subject ID in the target sheet (B5:B104)
        for target_row in range(5, 105):
            if target_sheet.cell(row=target_row, column=2).value == subject_id:
                print(f"Found Subject ID in {sheet_name} at row {target_row}")
                
                # Get the values from the source sheet (C5:AA5 to C13:AA13)
                row_offset = sheet_index + 4  # Mapping C5:AA5 -> q1_cat, C6:AA6 -> q2_cat, etc.
                source_values = [categorized_sheet.cell(row=row_offset, column=col).value
                                 for col in range(3, 28)]  # Columns C to AA
                print(f"Source values: {source_values}")

                # Paste the source values into the corresponding row (Columns C to AA) in the target sheet
                for col, value in enumerate(source_values, start=3):
                    target_sheet.cell(row=target_row, column=col, value=value)
                break  # Break once the subject ID match is found

    # Save the updated target workbook
    target_wb.save(target_file)
    print(f"Saved updated target workbook: {target_file}")

def process_source_folder(source_folder, target_file):
    # Iterate through all files in the source folder
    for filename in os.listdir(source_folder):
        if filename.endswith(".xlsx"):
            source_file = os.path.join(source_folder, filename)
            print(f"Processing file: {source_file}")
            copy_categorized_values(source_file, target_file)

# Example usage:
source_folder = "/projects/digimind/arna/fil"  # Replace with the path to your source folder
target_file = "/projects/digimind/arna/ALL_video_task_experience.xlsx"  # Replace with the path to your target file
process_source_folder(source_folder, target_file)
